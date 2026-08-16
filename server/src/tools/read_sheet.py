import sys, os
from openpyxl import load_workbook

# 把 .xlsx/.xls/.xlsm 抽取为可读文本（行×列，以 | 分隔）。
# 用法: python read_sheet.py <file.xlsx>
# 输出到 stdout；失败信息以 FAIL: 前缀打印到 stdout 并以非 0 退出。
MAX_ROWS_PER_SHEET = 500

def main():
    if len(sys.argv) < 2:
        print("FAIL: usage: read_sheet.py <file.xlsx>")
        sys.exit(2)
    src = sys.argv[1]
    if not os.path.exists(src):
        print("FAIL: FILE_NOT_FOUND: %s" % src)
        sys.exit(3)
    try:
        wb = load_workbook(src, data_only=True)
    except Exception as e:
        print("FAIL: OPEN_FAILED: %s" % e)
        sys.exit(4)

    out = []
    out.append("=== %s ===" % os.path.basename(src))
    out.append("工作表: %s" % ", ".join(wb.sheetnames))
    for sn in wb.sheetnames:
        ws = wb[sn]
        out.append("=== 工作表: %s (%d行 x %d列) ===" % (sn, ws.max_row, ws.max_column))
        for r in range(1, min(ws.max_row, MAX_ROWS_PER_SHEET) + 1):
            vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                vals.append("" if v is None else str(v))
            line = " | ".join(vals).rstrip(" |")
            if line.strip():
                out.append(line)
        if ws.max_row > MAX_ROWS_PER_SHEET:
            out.append("... (仅显示前 %d 行，共 %d 行)" % (MAX_ROWS_PER_SHEET, ws.max_row))
    wb.close()
    sys.stdout.write("\n".join(out))

if __name__ == "__main__":
    main()
